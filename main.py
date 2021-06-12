from flask import Flask, render_template, request

from flask_mqtt import Mqtt

from openpyxl import Workbook, load_workbook, workbook


app = Flask(__name__)
app.config['MQTT_BROKER_URL'] = 'mqtt.flespi.io'
app.config['MQTT_BROKER_PORT'] = 1883
app.config['MQTT_USERNAME'] = 'FlespiToken vTsX8UaHjIoRDbemkaTkhGUHkHhNVd3eUk2Q20nbzBT2AVcJVTT0pDyx9D8SG0Wq'
app.config['MQTT_PASSWORD'] = ''
app.config['MQTT_REFRESH_TIME'] = 1.0

mqtt = Mqtt(app)

# Subscribe to topic
@mqtt.on_connect()
def handle_connect(client, userdata, flags, rc):
    print("CONNECTED: SUCCESFUL")
    mqtt.subscribe('topic/esp32/temperature')
    mqtt.subscribe('topic/huminity')

@mqtt.on_message()
def handle_mqtt_message(client, userdata, message):
    data = dict(
        topic=message.topic,
        payload=message.payload.decode()
    )
    stt = 0
    counter_row = 0
    if data['topic'] == "topic/esp32/temperature":
        database_load = load_workbook("data.xlsx")
        database= database_load.active
        database.cell(row = counter_row, column = 1,value = stt)
        database.cell(row = counter_row, column = 2,value = data['payload'])
        database.save('data.xlsx')
        print ("Data from topic/esp32/temperature is: ", data['payload'])
# Publish topic


@app.route('/')
def index():
    return render_template('index.html')

